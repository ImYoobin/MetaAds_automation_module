"""Execution orchestrator for standalone Meta export."""

from __future__ import annotations

import os
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from .engine.file_naming import format_name
from .engine.meta_automation import MetaAutomation
from .config import StandaloneMetaExportConfig
from .constants import REQUIRED_SHEETS
from .runtime import (
    build_sb_kwargs,
    configure_insecure_https_bootstrap,
    configure_logger,
    ensure_browser_driver_ready,
    load_embedded_engine_config,
    progress_log_cb,
    verify_download_context,
)
from .internal.url_builder import build_report_export_url, build_report_view_url


@dataclass(frozen=True)
class StandaloneRunResult:
    run_id: str
    raw_files_by_sheet: dict[str, str]
    unified_workbook_path: str
    unified_workbook_size_bytes: int
    unified_sheet_names: list[str]
    missing_columns_by_sheet: dict[str, list[str]]
    log_file: str
    activity_name: str = ""


@dataclass(frozen=True)
class StandaloneBatchRunResult:
    run_id: str
    activity_results: list[StandaloneRunResult]
    log_file: str


def _now_run_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _now_yymmdd() -> str:
    # Include time to avoid duplicate export names within the same date.
    # Duplicate names can make history-row matching flaky in Meta Exports UI.
    return datetime.now().strftime("%y%m%d_%H%M%S")


def _sheet_display_name(sheet_key: str, display_map: dict[str, str]) -> str:
    return str(display_map.get(sheet_key) or sheet_key)


def _normalize_engine_browser(browser: str) -> str:
    normalized = str(browser or "").strip().lower()
    if normalized == "msedge":
        return "edge"
    if normalized == "chrome":
        return "chrome"
    raise RuntimeError(f"Unsupported browser: {browser} (allowed: msedge, chrome)")


def _read_workbook_sheet_names(payload: bytes) -> list[str]:
    try:
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        return list(wb.sheetnames)
    except Exception:
        return []


def _load_sheet_source_dataframes(
    *,
    raw_files_by_sheet: dict[str, str],
    parse_payload_to_df: Any,
    logger,
) -> dict[str, Any]:
    source_df_by_sheet: dict[str, Any] = {}
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    for sheet_key, file_path in raw_files_by_sheet.items():
        payload = Path(file_path).read_bytes()
        df = parse_payload_to_df(
            payload=payload,
            content_type=content_type,
            sheet_key=sheet_key,
        )
        source_df_by_sheet[sheet_key] = df
        logger.info(
            "source_df_loaded sheet=%s rows=%s cols=%s path=%s",
            sheet_key,
            int(getattr(df, "shape", [0, 0])[0]),
            int(getattr(df, "shape", [0, 0])[1]),
            file_path,
        )
    return source_df_by_sheet


def _load_transformer_components() -> tuple[dict[str, str], Any, Any]:
    try:
        from .transformer import (
            META_SHEET_KEY_TO_DISPLAY,
            MetaExportTransformer,
            parse_meta_export_payload_to_dataframe,
        )
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Missing dependency for standalone transform. "
            "Install required packages: `pip install pandas openpyxl`."
        ) from exc

    return (
        dict(META_SHEET_KEY_TO_DISPLAY),
        MetaExportTransformer,
        parse_meta_export_payload_to_dataframe,
    )


def _export_raw_files_for_activity(
    *,
    meta: Any,
    sb: Any,
    config: StandaloneMetaExportConfig,
    display_map: dict[str, str],
    logger,
    yymmdd: str,
) -> dict[str, str]:
    raw_files_by_sheet: dict[str, str] = {}
    enabled_sheet_keys = [
        sheet_key
        for sheet_key in REQUIRED_SHEETS
        if sheet_key in config.sheet_config_by_key and config.sheet_config_by_key[sheet_key].enabled
    ]
    if not enabled_sheet_keys:
        raise RuntimeError(
            f"No enabled required sheets configured for activity `{config.activity_name}`."
        )

    for sheet_key in enabled_sheet_keys:
        sheet_cfg = config.sheet_config_by_key[sheet_key]
        display_sheet_name = _sheet_display_name(sheet_key, display_map)
        report_to_sheet = {display_sheet_name: display_sheet_name}
        brand_cfg = {
            "brand_ko": config.brand_name,
            "meta_act_id": sheet_cfg.act_id,
            "meta_ad_account_id": sheet_cfg.act_id,
            "meta_business_id": sheet_cfg.business_id,
            "meta_global_scope_id": sheet_cfg.global_scope_id or sheet_cfg.business_id,
            "meta_view_event_source": config.view_event_source,
            "meta_export_event_source": config.export_event_source,
        }

        view_url = build_report_view_url(
            act_id=sheet_cfg.act_id,
            business_id=sheet_cfg.business_id,
            global_scope_id=sheet_cfg.global_scope_id,
            report_id=sheet_cfg.report_id,
            event_source=config.view_event_source,
        )
        export_url = build_report_export_url(
            act_id=sheet_cfg.act_id,
            business_id=sheet_cfg.business_id,
            global_scope_id=sheet_cfg.global_scope_id,
            event_source=config.export_event_source,
        )
        logger.info(
            "sheet_url_snapshot activity=%s sheet=%s report_id=%s view_url=%s export_url=%s",
            config.activity_name,
            sheet_key,
            sheet_cfg.report_id,
            view_url,
            export_url,
        )

        file_path = meta._export_report_via_view_id(  # noqa: SLF001 - reuse production-tested internal flow
            sb=sb,
            brand_cfg=brand_cfg,
            report_name=display_sheet_name,
            report_id=sheet_cfg.report_id,
            activity_for_filename=config.activity_name,
            yymmdd=yymmdd,
            report_to_sheet=report_to_sheet,
            sheet_key=sheet_key,
        )
        size_bytes = os.path.getsize(file_path)
        raw_files_by_sheet[sheet_key] = file_path
        logger.info(
            "sheet_downloaded activity=%s sheet=%s display=%s path=%s size_bytes=%s",
            config.activity_name,
            sheet_key,
            display_sheet_name,
            file_path,
            size_bytes,
        )
    return raw_files_by_sheet


def _build_activity_result(
    *,
    run_id: str,
    log_file: str,
    config: StandaloneMetaExportConfig,
    raw_files_by_sheet: dict[str, str],
    yymmdd: str,
    output_dir: Path,
    naming_config: dict[str, Any],
    format_name: Any,
    parse_payload_to_df: Any,
    transformer: Any,
    logger,
) -> StandaloneRunResult:
    source_df_by_sheet = _load_sheet_source_dataframes(
        raw_files_by_sheet=raw_files_by_sheet,
        parse_payload_to_df=parse_payload_to_df,
        logger=logger,
    )

    workbook_bytes, missing_columns_by_sheet = transformer.build_unified_workbook(source_df_by_sheet)
    final_pattern = str(
        naming_config.get("final_file_name_pattern") or "{brand}_{activity}_{yyMMdd}_{sheet}.xlsx"
    )
    output_filename = format_name(
        pattern=final_pattern,
        brand=config.brand_name,
        activity=config.activity_name,
        yymmdd=yymmdd,
        sheet="Unified",
    )
    output_path = (output_dir / output_filename).resolve()
    output_path.write_bytes(workbook_bytes)

    unified_sheet_names = _read_workbook_sheet_names(workbook_bytes)
    logger.info(
        "unified_workbook_created activity=%s path=%s size_bytes=%s sheet_names=%s",
        config.activity_name,
        output_path,
        len(workbook_bytes),
        unified_sheet_names,
    )
    logger.info("missing_columns_by_sheet activity=%s value=%s", config.activity_name, missing_columns_by_sheet)
    return StandaloneRunResult(
        run_id=run_id,
        raw_files_by_sheet=raw_files_by_sheet,
        unified_workbook_path=str(output_path),
        unified_workbook_size_bytes=len(workbook_bytes),
        unified_sheet_names=unified_sheet_names,
        missing_columns_by_sheet=missing_columns_by_sheet,
        log_file=log_file,
        activity_name=config.activity_name,
    )


def run_standalone_export_batch(
    *,
    configs: Sequence[StandaloneMetaExportConfig],
    config_path: Path,
    output_dir: Path,
    downloads_dir: Path,
    logs_dir: Path,
    browser: str,
) -> StandaloneBatchRunResult:
    config_list = list(configs)
    if not config_list:
        raise ValueError("At least one activity config is required.")

    run_id = _now_run_id()
    logger = configure_logger(logs_dir=logs_dir, run_id=run_id)
    log_file = str((logs_dir / f"run_{run_id}.log").resolve())
    engine_browser = _normalize_engine_browser(browser)
    mvp_config = load_embedded_engine_config()
    display_map, TransformerClass, parse_meta_export_payload_to_dataframe = _load_transformer_components()
    transformer = TransformerClass()

    downloads_dir = downloads_dir.resolve()
    output_dir = output_dir.resolve()
    downloads_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    activity_names = [cfg.activity_name for cfg in config_list]
    logger.info(
        "standalone_batch_start config=%s engine=embedded browser=%s engine_browser=%s activity_count=%s activity_names=%s",
        str(config_path.resolve()),
        browser,
        engine_browser,
        len(config_list),
        activity_names,
    )

    meta_config = dict(mvp_config.get("meta") or {})
    naming_config = dict(mvp_config.get("naming") or {})
    meta_config["home_url"] = "https://business.facebook.com/"

    configure_insecure_https_bootstrap()
    meta = MetaAutomation(
        meta_config=meta_config,
        naming_config=naming_config,
        download_dir=str(downloads_dir),
        headless=False,
    )

    try:
        from seleniumbase import SB
    except Exception as exc:
        raise RuntimeError("seleniumbase is required for standalone export.") from exc

    sb_kwargs = build_sb_kwargs(meta, engine_browser)
    logger.info("browser_launch_kwargs=%s", {k: v for k, v in sb_kwargs.items() if k != "proxy"})
    ensure_browser_driver_ready(browser=engine_browser, logger=logger)

    activity_results: list[StandaloneRunResult] = []
    try:
        with SB(**sb_kwargs) as sb:
            meta._enable_download_behavior(sb)  # noqa: SLF001 - reusing stable internal hook
            verify_download_context(meta, watcher_dir=downloads_dir, logger=logger)

            home_url = "https://business.facebook.com/"
            logger.info("open_meta_home url=%s", home_url)
            sb.open(home_url)

            logger.info("wait_manual_login start")
            meta._wait_for_meta_login(sb, progress_cb=progress_log_cb(logger, "login_wait"))  # noqa: SLF001
            logger.info("wait_manual_login done current_url=%s", sb.get_current_url())

            for config in config_list:
                logger.info(
                    "activity_start brand=%s/%s activity=%s selection_mode=%s account_groups=%s",
                    config.brand_code,
                    config.brand_name,
                    config.activity_name,
                    config.selection_mode,
                    config.account_groups,
                )
                yymmdd = _now_yymmdd()
                raw_files_by_sheet = _export_raw_files_for_activity(
                    meta=meta,
                    sb=sb,
                    config=config,
                    display_map=display_map,
                    logger=logger,
                    yymmdd=yymmdd,
                )
                activity_result = _build_activity_result(
                    run_id=run_id,
                    log_file=log_file,
                    config=config,
                    raw_files_by_sheet=raw_files_by_sheet,
                    yymmdd=yymmdd,
                    output_dir=output_dir,
                    naming_config=naming_config,
                    format_name=format_name,
                    parse_payload_to_df=parse_meta_export_payload_to_dataframe,
                    transformer=transformer,
                    logger=logger,
                )
                activity_results.append(activity_result)
                logger.info(
                    "activity_done activity=%s workbook=%s",
                    config.activity_name,
                    activity_result.unified_workbook_path,
                )
    except Exception:
        logger.exception("browser_flow_failed run_id=%s", run_id)
        raise

    logger.info("standalone_batch_done run_id=%s activity_count=%s", run_id, len(activity_results))
    return StandaloneBatchRunResult(
        run_id=run_id,
        activity_results=activity_results,
        log_file=log_file,
    )


def run_standalone_export(
    *,
    config: StandaloneMetaExportConfig,
    config_path: Path,
    output_dir: Path,
    downloads_dir: Path,
    logs_dir: Path,
    browser: str,
) -> StandaloneRunResult:
    batch_result = run_standalone_export_batch(
        configs=[config],
        config_path=config_path,
        output_dir=output_dir,
        downloads_dir=downloads_dir,
        logs_dir=logs_dir,
        browser=browser,
    )
    if not batch_result.activity_results:
        raise RuntimeError("No export result generated for the requested activity.")
    return batch_result.activity_results[0]
