"""Standalone Meta export parsing + workbook transformation helpers."""

from __future__ import annotations

import io
import json
import logging
import re
from collections.abc import Mapping
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, time
from typing import Any

import pandas as pd
from openpyxl import load_workbook


META_SHEET_KEY_TO_DISPLAY: dict[str, str] = {
    "overall": "Overall",
    "demo": "Demo",
    "overall_bof": "Overall-BoF",
    "demo_bof": "Demo-BoF",
    "time": "Time",
    "time_bof": "Time-BoF",
}

META_SOURCE_WORKSHEET_BY_KEY: dict[str, str] = {
    "overall": "Raw Data Report",
    "demo": "Creative Reporting",
    "time": "Raw Data Report",
    "overall_bof": "Raw Data Report",
    "demo_bof": "Creative Reporting",
    "time_bof": "Raw Data Report",
}

TARGET_TO_SOURCE_COLUMN_MAP: dict[str, str] = {
    "CPQV (KRW)": "Cost per QualifiedVisit",
}

_LOGGER = logging.getLogger(__name__)
_SCIENTIFIC_NOTATION_PATTERN = re.compile(r"^[+-]?(?:\d+\.?\d*|\.\d+)[eE][+-]?\d+$")
_ID_COLUMN_KEYS = {"campaignid", "adsetid", "adid"}


@dataclass(frozen=True)
class TemplateSheetConfig:
    sheet_name: str
    columns: list[str]


_META_TEMPLATE_SHEETS: tuple[TemplateSheetConfig, ...] = (
    TemplateSheetConfig(
        "Overall",
        [
            "Placement",
            "Day",
            "Campaign ID",
            "Ad set ID",
            "Ad ID",
            "Device platform",
            "Campaign name",
            "Ad set name",
            "Ad name",
            "Reach",
            "Impressions",
            "Frequency",
            "Attribution setting",
            "Amount spent (KRW)",
            "Link clicks",
            "CPQV (KRW)",
            "Clicks (all)",
            "Video plays at 50%",
            "Video plays at 100%",
            "call_to_action_type",
            "Objective",
            "Views",
            "Campaign Budget",
            "Campaign Budget Type",
            "QualifiedVisit",
            "Purchases",
            "3-second video plays",
            "Buying type",
            "Bid",
            "Bid type",
            "Performance goal",
            "Included custom audiences",
            "Ad Set Budget",
            "Ad Set Budget Type",
            "Reporting starts",
            "Reporting ends",
            "ThruPlays",
        ],
    ),
    TemplateSheetConfig(
        "Overall-BoF",
        [
            "Placement",
            "Day",
            "Campaign ID",
            "Ad set ID",
            "Ad ID",
            "Device platform",
            "Campaign name",
            "Ad set name",
            "Ad name",
            "Reach",
            "Impressions",
            "Frequency",
            "Attribution setting",
            "Amount spent (KRW)",
            "Link clicks",
            "CPQV (KRW)",
            "Clicks (all)",
            "Video plays at 50%",
            "Video plays at 100%",
            "call_to_action_type",
            "Objective",
            "Views",
            "Campaign Budget",
            "Campaign Budget Type",
            "3-second video plays",
            "Buying type",
            "Bid",
            "Bid type",
            "Performance goal",
            "Included custom audiences",
            "Ad Set Budget",
            "Ad Set Budget Type",
            "Result type",
            "Results",
            "Reporting starts",
            "Reporting ends",
            "ThruPlays",
        ],
    ),
    TemplateSheetConfig(
        "Demo",
        [
            "Day",
            "Campaign ID",
            "Ad set ID",
            "Ad ID",
            "Campaign name",
            "Ad set name",
            "Ad name",
            "Age",
            "Gender",
            "Reach",
            "Impressions",
            "Frequency",
            "Attribution setting",
            "Amount spent (KRW)",
            "Link clicks",
            "CPQV (KRW)",
            "Clicks (all)",
            "Video plays at 50%",
            "Video plays at 100%",
            "call_to_action_type",
            "Objective",
            "Views",
            "Campaign Budget",
            "Campaign Budget Type",
            "QualifiedVisit",
            "Purchases",
            "3-second video plays",
            "Buying type",
            "Bid",
            "Bid type",
            "Performance goal",
            "Included custom audiences",
            "Ad Set Budget",
            "Ad Set Budget Type",
            "Reporting starts",
            "Reporting ends",
            "ThruPlays",
        ],
    ),
    TemplateSheetConfig(
        "Demo-BoF",
        [
            "Day",
            "Campaign ID",
            "Ad set ID",
            "Ad ID",
            "Campaign name",
            "Ad set name",
            "Ad name",
            "Age",
            "Gender",
            "Reach",
            "Impressions",
            "Frequency",
            "Amount spent (KRW)",
            "Link clicks",
            "Clicks (all)",
            "Video plays at 50%",
            "Video plays at 100%",
            "call_to_action_type",
            "Objective",
            "Views",
            "Campaign Budget",
            "Campaign Budget Type",
            "Purchases",
            "3-second video plays",
            "Buying type",
            "Bid",
            "Bid type",
            "Included custom audiences",
            "Ad Set Budget",
            "Ad Set Budget Type",
            "Result type",
            "Results",
            "Reporting starts",
            "Reporting ends",
            "ThruPlays",
        ],
    ),
    TemplateSheetConfig(
        "Time",
        [
            "Day",
            "Campaign ID",
            "Ad set ID",
            "Ad ID",
            "Campaign name",
            "Ad set name",
            "Ad name",
            "Time of day (viewer's time zone)",
            "Reach",
            "Impressions",
            "Frequency",
            "Attribution setting",
            "Amount spent (KRW)",
            "Link clicks",
            "Clicks (all)",
            "Video plays at 50%",
            "Video plays at 100%",
            "call_to_action_type",
            "Objective",
            "Views",
            "Campaign Budget",
            "Campaign Budget Type",
            "QualifiedVisit",
            "Purchases",
            "3-second video plays",
            "Buying type",
            "Bid",
            "Bid type",
            "Performance goal",
            "Included custom audiences",
            "Ad Set Budget",
            "Ad Set Budget Type",
            "CPQV (KRW)",
            "Reporting starts",
            "Reporting ends",
            "ThruPlays",
        ],
    ),
    TemplateSheetConfig(
        "Time-BoF",
        [
            "Day",
            "Campaign ID",
            "Ad set ID",
            "Ad ID",
            "Campaign name",
            "Ad set name",
            "Ad name",
            "Time of day (viewer's time zone)",
            "Reach",
            "Impressions",
            "Frequency",
            "Amount spent (KRW)",
            "Link clicks",
            "Clicks (all)",
            "Video plays at 50%",
            "Video plays at 100%",
            "call_to_action_type",
            "Objective",
            "Views",
            "Campaign Budget",
            "Campaign Budget Type",
            "3-second video plays",
            "Buying type",
            "Bid",
            "Bid type",
            "Included custom audiences",
            "Ad Set Budget",
            "Ad Set Budget Type",
            "Reporting starts",
            "Reporting ends",
            "ThruPlays",
        ],
    ),
)


def parse_meta_export_payload_to_dataframe(
    *,
    payload: bytes,
    content_type: str,
    sheet_key: str,
) -> pd.DataFrame:
    normalized_key = _normalize_sheet_key(sheet_key)
    lower_content_type = str(content_type or "").lower()

    if _looks_like_xlsx(payload, lower_content_type):
        return _read_xlsx_payload(payload, normalized_key)
    if "json" in lower_content_type:
        return _read_json_payload(payload)
    return _read_csv_payload(payload)


def _normalize_sheet_key(sheet_key: str) -> str:
    return "_".join(str(sheet_key or "").strip().lower().replace("-", "_").split())


def _looks_like_xlsx(payload: bytes, content_type: str) -> bool:
    if "spreadsheetml" in content_type or "ms-excel" in content_type:
        return True
    return payload[:2] == b"PK"


def _normalize_header_value(raw_value: Any, index: int) -> str:
    text = str(raw_value).strip() if raw_value is not None else ""
    return text if text else f"Column_{index}"


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: set[str] = set()
    duplicate_roots: list[str] = []
    out: list[str] = []

    for header in headers:
        candidate = str(header or "").strip()
        if candidate not in seen:
            seen.add(candidate)
            out.append(candidate)
            continue

        if candidate not in duplicate_roots:
            duplicate_roots.append(candidate)
        suffix = 2
        while f"{candidate}__dup{suffix}" in seen:
            suffix += 1
        unique_name = f"{candidate}__dup{suffix}"
        seen.add(unique_name)
        out.append(unique_name)

    if duplicate_roots:
        _LOGGER.warning("duplicate_headers_detected names=%s", duplicate_roots)
    return out


def _normalize_id_text(text: str) -> str:
    value = str(text or "").strip()
    if not value:
        return ""
    if value.startswith("+"):
        value = value[1:]
    # Keep plain digits as-is for performance and exactness.
    if value.isdigit():
        return value

    try:
        decimal_value = Decimal(value)
    except InvalidOperation:
        return value

    if decimal_value.is_nan() or decimal_value.is_infinite():
        return value

    integral = decimal_value.to_integral_value()
    if decimal_value == integral:
        normalized = format(integral, "f")
    else:
        normalized = format(decimal_value.normalize(), "f").rstrip("0").rstrip(".")
    return normalized or value


def _value_to_trimmed_text(
    *,
    value: Any,
    column_name: str,
    row_idx: int,
) -> str:
    if value is None:
        return ""

    if isinstance(value, str):
        text = value.strip()
    elif isinstance(value, datetime):
        text = value.isoformat(sep=" ")
    elif isinstance(value, (date, time)):
        text = value.isoformat()
    else:
        text = str(value).strip()

    if text and _SCIENTIFIC_NOTATION_PATTERN.match(text):
        _LOGGER.warning(
            "scientific_notation_detected column=%s row=%s value=%s",
            column_name,
            row_idx,
            text,
        )

    if _is_id_column(column_name):
        normalized_id = _normalize_id_text(text)
        if normalized_id != text:
            _LOGGER.info(
                "id_normalized column=%s row=%s before=%s after=%s",
                column_name,
                row_idx,
                text,
                normalized_id,
            )
        return normalized_id
    return text


def _read_xlsx_payload(payload: bytes, sheet_key: str) -> pd.DataFrame:
    expected_sheet = META_SOURCE_WORKSHEET_BY_KEY.get(sheet_key, "Raw Data Report")
    workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    try:
        available = list(workbook.sheetnames)
        sheet_name = expected_sheet if expected_sheet in available else (available[0] if available else None)
        if not sheet_name:
            return pd.DataFrame()

        worksheet = workbook[sheet_name]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return pd.DataFrame()

        headers = _dedupe_headers([
            _normalize_header_value(raw_value, index=idx)
            for idx, raw_value in enumerate(header_row, start=1)
        ])

        records: list[list[str]] = []
        for row_idx, row in enumerate(rows_iter, start=2):
            row_values: list[str] = []
            for col_idx, column_name in enumerate(headers, start=1):
                raw_value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                text = _value_to_trimmed_text(
                    value=raw_value,
                    column_name=column_name,
                    row_idx=row_idx,
                )
                row_values.append(text)

            # Remove fully empty rows only (minimal cleanup policy).
            if all(value == "" for value in row_values):
                continue
            records.append(row_values)

        return pd.DataFrame(records, columns=headers, dtype=object)
    finally:
        workbook.close()


def _read_csv_payload(payload: bytes) -> pd.DataFrame:
    text = payload.decode("utf-8-sig", errors="replace")
    best_df: pd.DataFrame | None = None
    best_score = -1
    for skip_rows in (0, 1, 2):
        try:
            df = pd.read_csv(io.StringIO(text), dtype=object, skiprows=skip_rows).fillna("")
        except Exception:
            continue
        score = len(df.columns)
        if score > best_score:
            best_df = df
            best_score = score
    return best_df if best_df is not None else pd.DataFrame()


def _read_json_payload(payload: bytes) -> pd.DataFrame:
    try:
        parsed = json.loads(payload.decode("utf-8", errors="replace"))
    except json.JSONDecodeError:
        return pd.DataFrame()

    rows = _json_rows(parsed)
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).fillna("")


def _json_rows(parsed: Any) -> list[dict[str, Any]]:
    if isinstance(parsed, list):
        return [item for item in parsed if isinstance(item, dict)]
    if isinstance(parsed, dict):
        for key in ("data", "rows", "results", "items"):
            value = parsed.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
    return []


class _TemplateBuilder:
    def build_meta_template(self) -> bytes:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for config in _META_TEMPLATE_SHEETS:
                df = pd.DataFrame(columns=config.columns)
                df.loc[0] = [""] * len(config.columns)
                df.to_excel(writer, sheet_name=config.sheet_name[:31], index=False)
        buffer.seek(0)
        return buffer.read()


class MetaExportTransformer:
    def __init__(self) -> None:
        self._template_builder = _TemplateBuilder()

    def build_unified_workbook(
        self,
        source_df_by_sheet_key: Mapping[str, pd.DataFrame],
    ) -> tuple[bytes, dict[str, list[str]]]:
        template_bytes = self._template_builder.build_meta_template()
        workbook = load_workbook(io.BytesIO(template_bytes))

        missing_by_sheet: dict[str, list[str]] = {}
        for sheet_key, target_sheet_name in META_SHEET_KEY_TO_DISPLAY.items():
            source_df = source_df_by_sheet_key.get(sheet_key, pd.DataFrame()).fillna("")
            target_columns = _template_header_columns(workbook[target_sheet_name])
            target_df, missing = _build_target_sheet_df(
                source_df=source_df,
                target_columns=target_columns,
                target_to_source_col=TARGET_TO_SOURCE_COLUMN_MAP,
                sheet_name=target_sheet_name,
            )
            target_df, _ = _drop_summary_rows(
                sheet_name=target_sheet_name,
                df=target_df,
                stage="before_sheet_write",
            )
            _write_df_to_sheet(workbook[target_sheet_name], target_df)
            missing_by_sheet[target_sheet_name] = missing

        out = io.BytesIO()
        workbook.save(out)
        out.seek(0)
        return out.read(), missing_by_sheet


def _normalize_day(series: pd.Series) -> pd.Series:
    if series.empty:
        return series
    if pd.api.types.is_datetime64_any_dtype(series):
        return series.dt.strftime("%Y-%m-%d")

    as_num = pd.to_numeric(series, errors="coerce")
    if as_num.notna().sum() > 0:
        as_dt = pd.to_datetime(as_num, unit="D", origin="1899-12-30", errors="coerce")
        if as_dt.notna().sum() > 0:
            return as_dt.dt.strftime("%Y-%m-%d").fillna(series.astype(str))

    as_dt = pd.to_datetime(series, errors="coerce")
    if as_dt.notna().sum() > 0:
        return as_dt.dt.strftime("%Y-%m-%d").fillna(series.astype(str))
    return series.astype(str)


def _build_target_sheet_df(
    *,
    source_df: pd.DataFrame,
    target_columns: list[str],
    target_to_source_col: Mapping[str, str],
    sheet_name: str = "",
) -> tuple[pd.DataFrame, list[str]]:
    out = pd.DataFrame(index=source_df.index.copy())
    missing: list[str] = []
    source_lookup = _build_source_col_lookup(source_df.columns)

    for target_col in target_columns:
        mapped_source_col = target_to_source_col.get(target_col, target_col)
        resolved_source_col = _resolve_source_col_name(
            source_lookup=source_lookup,
            candidate_names=[mapped_source_col, target_col],
        )
        if resolved_source_col:
            out[target_col] = source_df[resolved_source_col]
        else:
            out[target_col] = ""
            missing.append(target_col)

    for id_col in ("Campaign ID", "Ad set ID", "Ad Set ID", "Ad ID"):
        if id_col in out.columns:
            out[id_col] = out[id_col].map(
                lambda value: _normalize_id_text(str(value).strip()) if not pd.isna(value) else ""
            )

    if "Day" in out.columns:
        out["Day"] = _normalize_day(out["Day"])
    out, _ = _drop_summary_rows(
        sheet_name=sheet_name or "(unknown)",
        df=out,
        stage="after_mapping",
    )
    return out, missing


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def _is_numeric_like_value(value: Any) -> bool:
    if _is_blank_value(value):
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float, Decimal)):
        return True

    text = str(value).strip()
    if not text:
        return False
    text = text.replace(",", "").replace("%", "")
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1].strip()}"
    if text.startswith("+"):
        text = text[1:]
    try:
        decimal_value = Decimal(text)
    except InvalidOperation:
        return False
    return not decimal_value.is_nan() and not decimal_value.is_infinite()


def _drop_summary_rows(
    *,
    sheet_name: str,
    df: pd.DataFrame,
    stage: str,
) -> tuple[pd.DataFrame, int]:
    if df.empty:
        return df, 0

    key_candidates = [
        "Day",
        "Campaign ID",
        "Ad set ID",
        "Ad Set ID",
        "Ad ID",
        "Campaign name",
        "Campaign Name",
        "Ad set name",
        "Ad Set Name",
        "Ad name",
        "Ad Name",
    ]
    key_cols = [col for col in key_candidates if col in df.columns]
    if not key_cols:
        return df, 0

    metric_cols = [col for col in df.columns if col not in key_cols]
    if not metric_cols:
        return df, 0

    drop_indexes: list[Any] = []
    for row_index, row in df.iterrows():
        if any(not _is_blank_value(row.get(col)) for col in key_cols):
            continue
        metric_values = [row.get(col) for col in metric_cols]
        if not any(not _is_blank_value(value) for value in metric_values):
            continue
        if not any(_is_numeric_like_value(value) for value in metric_values):
            continue
        drop_indexes.append(row_index)

    removed = len(drop_indexes)
    if removed <= 0:
        return df, 0

    _LOGGER.info(
        "summary_row_removed sheet=%s stage=%s removed=%s",
        sheet_name,
        stage,
        removed,
    )
    return df.drop(index=drop_indexes).reset_index(drop=True), removed


def _normalize_col_name(value: str) -> str:
    return "".join(str(value or "").strip().lower().split())


def _is_id_column(column_name: str) -> bool:
    return _normalize_col_name(column_name) in _ID_COLUMN_KEYS


def _build_source_col_lookup(columns: Any) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for col in columns:
        key = _normalize_col_name(str(col))
        if key and key not in lookup:
            lookup[key] = str(col)
    return lookup


def _resolve_source_col_name(
    *,
    source_lookup: Mapping[str, str],
    candidate_names: list[str],
) -> str:
    for name in candidate_names:
        key = _normalize_col_name(name)
        if key in source_lookup:
            return source_lookup[key]
    return ""


def _template_header_columns(ws) -> list[str]:
    headers: list[str] = []
    if ws.max_column <= 0:
        return headers
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        text = str(value).strip() if value is not None else ""
        if text:
            headers.append(text)
    return headers


def _clear_sheet_values(ws) -> None:
    if ws.max_row <= 0 or ws.max_column <= 0:
        return
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None


def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    _clear_sheet_values(ws)
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    if df.empty:
        return
    id_column_flags = [_is_id_column(str(col_name)) for col_name in df.columns]
    for row_idx, row_values in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if id_column_flags[col_idx - 1]:
                text_value = "" if pd.isna(value) else str(value).strip()
                cell.value = text_value
                cell.number_format = "@"
            else:
                cell.value = "" if pd.isna(value) else value
